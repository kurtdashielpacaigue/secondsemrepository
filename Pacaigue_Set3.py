import tkinter as tk
from tkinter import ttk,messagebox
import openpyxl as op

def display():
    workbook = op.load_workbook("orderDB.xlsx")
    sheet = workbook.active

    for row in table.get_children():
        table.delete(row)

    for row in sheet.iter_rows(min_row=2,values_only=True):
        table.insert("",tk.END,values=row)





def input_validation():
    cna=cname_entry.get()
    prod=product_entry.get()
    qt=qty_entry.get()
    pri= price_entry.get()

    if not cna  or not prod  :
        messagebox.showerror("Error","must be a word  ")
        return False
    
    if not pri.isdigit or not qt.isdigit:
        messagebox.showerror("Error","price must be a number")
        return False
    return True

def saving():
    if not input_validation():
        return False
    
    cna=cname_entry.get()
    prod=product_entry.get()
    qt=int(qty_entry.get())
    pri=int( price_entry.get())

    total = qt *pri

    workbook = op.load_workbook("orderDB.xlsx")
    sheet=workbook.active



    totals= sheet.max_row
    sheet.append([totals,prod,cna,qt,pri,total])
    workbook.save("orderDB.xlsx")
    messagebox.showinfo("Success","Record added")
    display()
    
    


def update():
    selected = table.focus()

    if not selected:
        messagebox.showerror("Error","select ")

    if not input_validation:
        return
    
    values = table.item(selected,"values")
    record_id = values[0]

    cna=cname_entry.get()
    prod=product_entry.get()
    qt=int(qty_entry.get())
    pri=int( price_entry.get())

    total = qt *pri
    

  

    workbook = op.load_workbook("excelDB.xlsx")
    sheet = workbook.active

    for rows in sheet.iter_rows(min_row=2):
        if str(rows[0].value) == str(record_id):
            rows[1].value = cna
            rows[2].value = prod
            rows [3].value =qt
            rows [4].value =pri
            rows [5].value =total
    
    workbook.save("excelDB.xlsx")
    messagebox.showinfo("Success","record updated ")
    display()

    
def auto_populate(event):
    selected = table.focus()
    values = table.item(selected,"values")

    if values:
      
        cname_entry.delete(0,tk.END)
        product_entry.delete(0,tk.END)
        qty_entry.delete(0,tk.END)
        price_entry.delete(0,tk.END)


        cname_entry.insert(0,values[1])
        product_entry.insert(0,values[2])
        qty_entry.insert(0,values[3])
        price_entry.insert(0,values[4])
    
    

def delete():
    selected = table.focus()

    if not selected:
        messagebox.showerror("Error","select something bruhhh")

    values = table.item(selected,"values")
    record_id = values[0]

    confirm = messagebox.askyesnocancel("Confirm","Are you sure you want to delete this record?")
    
    if not confirm:
        return

    workbook = op.load_workbook("orderDB.xlsx")
    sheet = workbook.active

    for i,row in enumerate(sheet.iter_rows(min_row=2),start=2):
        if str(row[0].value) == str(record_id):
            sheet.delete_rows(i)
            break

    workbook.save("orderDB.xlsx")
    messagebox.showinfo("Success","Record deleted successfully")
    display()



   
window = tk.Tk()
window.title("Simple Ordering System")
window.configure(bg="lightblue")

# Form Title
title = tk.Label(window, text="Simple Ordering System", font=("Times New Roman", 14, "bold"), bg="lightblue")
title.grid(row=0, column=0, columnspan=6)

# Frame
genframe = tk.Frame(window, bg="lightblue", bd=2, relief="groove")
genframe.grid(row=1, column=0, columnspan=7, padx=10, pady=10)

# Customer Name Entry
cname_entry = tk.Entry(genframe, font=("Poppins", 12))
cname_entry.grid(row=2, column=1, columnspan=2, padx=10, pady=(10, 0))

cname_label = tk.Label(genframe, text="Customer Name", font=("Poppins", 10, "italic"), bg="lightblue")
cname_label.grid(row=3, column=1, columnspan=2)

# Product Entry
product_entry = tk.Entry(genframe, font=("Poppins", 12))
product_entry.grid(row=2, column=3, columnspan=2, padx=10, pady=(10, 0))

product_label = tk.Label(genframe, text="Product", font=("Poppins", 10, "italic"), bg="lightblue")
product_label.grid(row=3, column=3, columnspan=2)

# Quantity Entry
qty_entry = tk.Entry(genframe, font=("Poppins", 12))
qty_entry.grid(row=4, column=1, columnspan=2, padx=10, pady=(10, 0))

qty_label = tk.Label(genframe, text="Quantity", font=("Poppins", 10, "italic"), bg="lightblue")
qty_label.grid(row=5, column=1, columnspan=2)

# Price Entry
price_entry = tk.Entry(genframe, font=("Poppins", 12))
price_entry.grid(row=4, column=3, columnspan=2, padx=10, pady=(10, 0))

price_label = tk.Label(genframe, text="Price", font=("Poppins", 10, "italic"), bg="lightblue")
price_label.grid(row=5, column=3, columnspan=2)

# Buttons
submit_btn = tk.Button(window, text="Submit", font=("Poppins", 12, "bold"), bg="lightpink",command=saving)
submit_btn.grid(row=6, column=1, pady=(10, 20))

update_btn = tk.Button(window, text="Update",font=("Poppins", 12, "bold"), bg="lightgreen",command=update)
update_btn.grid(row=6, column=2)

delete_btn = tk.Button(window, text="Delete", bg="red", fg="white",font=("Poppins", 12, "bold"),command=delete)
delete_btn.grid(row=6, column=3)

# Table
table = ttk.Treeview(
    window,
    columns=("Order ID", "Customer Name", "Product", "Quantity", "Price", "Total"),
    show="headings"
)

for headings in ("Order ID", "Customer Name", "Product", "Quantity", "Price", "Total"):
    table.heading(headings, text=headings)

table.grid(row=7, column=0, columnspan=6, padx=10, pady=10)

display()


window.mainloop()
