import time
import os
from openpyxl import Workbook, load_workbook

filename = "favorite_peoples1.xlsx"

if os.path.exists(filename):
    wb = load_workbook(filename)
    ws = wb.active

else:
    wb = Workbook()
    ws = wb.active

    headers = ["ID", "First Name", "Last Name", "Birth Year", "Age"]
    ws.append(headers)

def computeage(birthyear):
    age = 2026 - int(birthyear)
    return age

def appentoexcel(firstname, lastname, birthyear):
    newid = ws.max_row

    data = [
        newid,
        firstname,
        lastname,
        int(birthyear),
        computeage(birthyear)
    ]

    ws.append(data)
    wb.save(filename)

    print("Your new info has been added successfully!\n")

for i in range(3):

    print(f"\nPerson #{i + 1}")

    firstname = input("Input Your First Name: ")
    lastname = input("Input Your Last Name: ")

    while True:
        birthyear = input("Birth Year: ")

        if birthyear.isdigit() == False:
            print("You must input a valid birth year")

        else:
            break

    appentoexcel(firstname, lastname, birthyear)

print("All data has been saved!")

